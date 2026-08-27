import csv
import io
from pathlib import Path


def load_txt(path: str) -> str:
    return Path(path).read_text(encoding="utf-8")


def load_md(path: str) -> str:
    return Path(path).read_text(encoding="utf-8")


def load_csv(path: str) -> str:
    with open(path, encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return ""
    header = rows[0]
    lines = [",".join(header), "-" * 40]
    for row in rows[1:]:
        lines.append(" | ".join(f"{h}: {v}" for h, v in zip(header, row) if v))
    return "\n".join(lines)


async def load_pdf(path: str) -> str:
    """Extract text from PDF. Falls back to OCR for scanned/image-based pages."""
    import fitz  # pymupdf
    from config import settings

    ocr = None
    parts: list[str] = []
    scanned_count = 0

    doc = fitz.open(path)
    try:
        for page_num, page in enumerate(doc):
            text = page.get_text().strip()
            min_chars = getattr(settings, 'ocr_fallback_threshold', 20)
            if text and len(text) >= min_chars:
                parts.append(text)
            else:
                # Page has little or no text — likely a scanned image, try OCR
                scanned_count += 1
                ocr_text = await _ocr_page(page, page_num, ocr)
                if ocr_text:
                    parts.append(ocr_text)
    finally:
        doc.close()

    if scanned_count > 0 and not parts:
        # All pages were scanned — raise if nothing could be extracted
        if not parts:
            raise ValueError(
                f"PDF contains only scanned images and OCR could not extract text. "
                f"({scanned_count} pages). Install tesseract-ocr to enable OCR."
            )

    return "\n\n".join(parts)


async def _ocr_page(page, page_num: int, ocr) -> str:
    """Run OCR on a single PDF page. Returns extracted text or empty string."""
    import fitz
    from config import settings

    if ocr is None:
        from ocr.factory import create_ocr
        ocr = create_ocr()
        if ocr is None or not ocr.is_available():
            return ""

    # Render page to image at 200 DPI for good OCR quality
    mat = fitz.Matrix(200 / 72, 200 / 72)
    pix = page.get_pixmap(matrix=mat)
    from PIL import Image
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

    text = ocr.recognize(img)
    if text and len(text) > settings.ocr_min_text_length:
        return text
    return ""


async def load_docx(path: str) -> str:
    from docx import Document as DocxDocument
    doc = DocxDocument(path)
    parts = []
    for para in doc.paragraphs:
        if para.text.strip():
            parts.append(para.text)
    return "\n".join(parts)


async def load_xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    all_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [[str(cell.value) if cell.value is not None else "" for cell in row] for row in ws.iter_rows()]
        if not rows:
            continue
        all_parts.append(f"## Sheet: {sheet_name}")
        header = rows[0]
        all_parts.append(" | ".join(header))
        all_parts.append("-" * 40)
        for row in rows[1:]:
            all_parts.append(" | ".join(f"{h}: {v}" for h, v in zip(header, row) if v))
        all_parts.append("")
    wb.close()
    return "\n".join(all_parts)


_LOADERS = {
    ".txt": load_txt,
    ".md": load_md,
    ".csv": load_csv,
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".xlsx": load_xlsx,
}


async def load_file(path: str) -> str:
    ext = Path(path).suffix.lower()
    loader = _LOADERS.get(ext)
    if loader is None:
        raise ValueError(f"Unsupported file type: {ext}")
    if ext in (".pdf", ".docx", ".xlsx"):
        return await loader(path)
    return loader(path)
